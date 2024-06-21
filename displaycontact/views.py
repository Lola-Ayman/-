from django.shortcuts import render
import openpyxl

# Create your views here.

from openpyxl import load_workbook
def disp(request):
    wb = load_workbook(r'C:\Projects\emp\names.xlsx')
    ws =wb['emp']

    noofrows =int(ws['F1'].value)
    itmes=[]
    for i in range(2, noofrows+1):
        item=[]
        item.append(ws['A' + str(i)].value)
        item.append(ws['B' + str(i)].value)
        item.append(ws['C' + str(i)].value)
        item.append(ws['D' + str(i)].value)
        
        itmes.append(item)
        
    return render(request,'displaydata.html',{'itmes':itmes})
