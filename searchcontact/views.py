from django.shortcuts import render


# Create your views here.

from openpyxl import load_workbook
from .forms import SearchForm
def search(request):
    form = SearchForm()
    itmes=[]


    if request.method == 'POST':
        
        form = SearchForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data

            wb = load_workbook(r'C:\Projects\emp\names.xlsx')
            ws =wb['emp']

            noofrows =int(ws['F1'].value)
        
            for i in range(2, noofrows+1):
               if cd['name'] == ws['A' + str(i)].value:
                item=[]
                item.append(ws['A' + str(i)].value)
                item.append(ws['B' + str(i)].value)
                item.append(ws['C' + str(i)].value)
                item.append(ws['D' + str(i)].value)
        
                itmes.append(item)


        
        
    return render(request,'displaydata1.html',{'itmes':itmes , 'form':form})

        




   
